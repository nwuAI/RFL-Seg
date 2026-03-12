import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime


def save_metrics_to_excel(cfg, dices, hds, ious, accs, ses, sps):

    dice_mean = np.nanmean(dices * 100, axis=0)
    dice_std = np.nanstd(dices * 100, axis=0)
    hd_mean = np.nanmean(hds, axis=0)
    hd_std = np.nanstd(hds, axis=0)
    iou_mean = np.nanmean(ious * 100, axis=0)
    iou_std = np.nanstd(ious * 100, axis=0)
    acc_mean = np.nanmean(accs * 100, axis=0)
    acc_std = np.nanstd(accs * 100, axis=0)
    se_mean = np.nanmean(ses * 100, axis=0)
    se_std = np.nanstd(ses * 100, axis=0)
    sp_mean = np.nanmean(sps * 100, axis=0)
    sp_std = np.nanstd(sps * 100, axis=0)


    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data_row = [timestamp, cfg.load_model_dict_name]


    num_classes = len(dice_mean)
    for i in range(1, num_classes):
        data_row.extend([
            f"{dice_mean[i]:.2f}±{dice_std[i]:.2f}",
            f"{iou_mean[i]:.2f}±{iou_std[i]:.2f}",
            f"{acc_mean[i]:.2f}±{acc_std[i]:.2f}",
            f"{se_mean[i]:.2f}±{se_std[i]:.2f}",
            f"{sp_mean[i]:.2f}±{sp_std[i]:.2f}",
            f"{hd_mean[i]:.2f}±{hd_std[i]:.2f}"
        ])


    headers = ['Time', 'Model']
    for i in range(1, num_classes):
        headers.extend([
            f'Class{i}_Dice',
            f'Class{i}_IoU',
            f'Class{i}_Acc',
            f'Class{i}_SE',
            f'Class{i}_SP',
            f'Class{i}_HD'
        ])


    save_path = os.path.join(cfg.save_dir, "eval.xlsx")


    if os.path.exists(save_path):

        book = load_workbook(save_path)
        sheet = book.active


        sheet.append(data_row)
    else:

        df = pd.DataFrame(columns=headers)

        df.to_excel(save_path, index=False)
        book = load_workbook(save_path)
        sheet = book.active
        sheet.append(data_row)


        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))


        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border


        for i, header in enumerate(headers, 1):
            sheet.column_dimensions[get_column_letter(i)].width = 14


    data_font = Font(size=11)
    data_alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))


    best_fill_high = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 高值最优-绿色
    best_fill_low = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # 低值最优-橙色


    for col_idx in range(3, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        values = []


        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet[f"{col_letter}{row_idx}"].value
            if cell_value is not None and '±' in str(cell_value):
                values.append(float(str(cell_value).split('±')[0]))

        if not values:
            continue


        metric_name = sheet[f"{col_letter}1"].value
        if "_HD" in metric_name:
            best_value = min(values)
            best_fill = best_fill_low
        else:
            best_value = max(values)
            best_fill = best_fill_high


        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet[f"{col_letter}{row_idx}"]
            if cell.value is not None and '±' in str(cell.value):
                cell_value = float(str(cell.value).split('±')[0])
                if abs(cell_value - best_value) < 0.01:
                    cell.fill = best_fill


    book.save(save_path)
    print(f"saved to: {save_path}")
    return dice_mean, hd_mean, iou_mean, acc_mean, se_mean, sp_mean, dice_std, hd_std, iou_std, acc_std, se_std, sp_std